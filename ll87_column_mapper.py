"""
LL87 Column Mapper: 2018 → 2019
Maps column names from the 2018 LL87 data export to the closest equivalent
in the 2019 LL87 data export using Jaccard token similarity.

Usage:
    python ll87_column_mapper.py \
        --old LL87_Header_2018.xlsx \
        --new LL87_Header_2019.xlsx \
        --out LL87_Column_Mapping_2018_to_2019.xlsx

Requirements:
    pip install pandas openpyxl
"""

import argparse
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

THRESHOLD = 0.15          # Minimum Jaccard score to count as a match
LOW_CONF_THRESHOLD = 0.25 # Below this, the match is flagged as low confidence
DESC_WEIGHT = 0.6         # Weight applied to description-based scores
DESC_WORDS = 15           # How many words of description to use

STOPWORDS = {
    'the', 'a', 'an', 'of', 'in', 'for', 'to', 'is', 'and', 'or',
    'on', 'at', 'by', 'as', 'be', 'are', 'with', 'that', 'this', 'which', 'it'
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def tokenize(s: str) -> set:
    """Lowercase, strip non-alphanumeric, remove stopwords."""
    s = re.sub(r'[^a-z0-9]', ' ', str(s).lower())
    return {w for w in s.split() if w not in STOPWORDS and len(w) > 1}


def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def best_match(name18: str, names19: list, t19_list: list) -> tuple:
    """
    Find the best matching 2019 column for a given 2018 column name.

    Scoring strategy:
      - leaf tokens  : tokens from the last '_'-separated segment
      - leaf2 tokens : tokens from the last two segments
      - full tokens  : tokens from the entire column name
      - desc tokens  : tokens from the first N words of the 2019 description
                       (weighted by DESC_WEIGHT)

    Returns (best_match_name, best_match_desc, best_score).
    """
    parts = str(name18).split('_')
    lt  = tokenize(parts[-1])
    l2t = tokenize(' '.join(parts[-2:])) if len(parts) >= 2 else lt
    ft  = tokenize(name18)

    best_score, best_name, best_desc = 0.0, '', ''

    for i, (tn, td) in enumerate(t19_list):
        score = max(
            jaccard(lt,  tn),
            jaccard(l2t, tn),
            jaccard(ft,  tn),
            jaccard(lt,  td) * DESC_WEIGHT,
        )
        if score > best_score:
            best_score = score
            best_name  = names19[i]
            best_desc  = t19_list[i][1]   # description token set → not needed below

    # Retrieve plain description text
    return best_name, best_score


# ---------------------------------------------------------------------------
# Core mapping
# ---------------------------------------------------------------------------

def build_mapping(old_path: str, new_path: str) -> pd.DataFrame:
    df18 = pd.read_excel(old_path)
    df19 = pd.read_excel(new_path, sheet_name='Column Information')

    names18 = df18['Column Name'].tolist()
    names19 = df19['Column Name'].tolist()
    descs19 = df19['Column Description'].fillna('').tolist()

    # Pre-tokenize 2019 columns
    t19_list = [
        (tokenize(n), tokenize(' '.join(str(d).split()[:DESC_WORDS])))
        for n, d in zip(names19, descs19)
    ]

    desc_lookup = dict(zip(names19, descs19))

    rows = []
    for name18 in names18:
        parts = str(name18).split('_')
        lt  = tokenize(parts[-1])
        l2t = tokenize(' '.join(parts[-2:])) if len(parts) >= 2 else lt
        ft  = tokenize(name18)

        best_score, best_name = 0.0, ''
        for i, (tn, td) in enumerate(t19_list):
            score = max(
                jaccard(lt,  tn),
                jaccard(l2t, tn),
                jaccard(ft,  tn),
                jaccard(lt,  td) * DESC_WEIGHT,
            )
            if score > best_score:
                best_score = score
                best_name  = names19[i]

        if best_score >= THRESHOLD:
            rows.append((name18, best_name, desc_lookup.get(best_name, ''), round(best_score, 3)))
        else:
            rows.append((name18, '', '', round(best_score, 3)))

    return pd.DataFrame(rows, columns=[
        '2018 Column Name', '2019 Column Name', '2019 Column Description', 'Match Score'
    ])


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(df: pd.DataFrame, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Column Mapping"

    # --- Styles ---
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', start_color='1F4E79')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fill_matched_even = PatternFill('solid', start_color='F5F5F5')
    fill_matched_odd  = PatternFill('solid', start_color='FFFFFF')
    fill_low_conf     = PatternFill('solid', start_color='FFF9C4')
    fill_unmatched    = PatternFill('solid', start_color='FFEBEE')

    # --- Header row ---
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # --- Data rows ---
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        name18, name19, desc, score = row
        is_matched = name19 != ''
        is_low     = is_matched and score < LOW_CONF_THRESHOLD

        for col, val in enumerate([name18, name19, desc, score], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = Font(name='Arial', size=10)
            cell.border    = border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 3))

            if not is_matched:
                cell.fill = fill_unmatched
            elif is_low:
                cell.fill = fill_low_conf
            elif row_idx % 2 == 0:
                cell.fill = fill_matched_even
            else:
                cell.fill = fill_matched_odd

    # --- Column widths ---
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = 'A2'

    # --- Legend sheet ---
    ws2 = wb.create_sheet("Legend")
    legend_rows = [
        ("Color",             "Meaning"),
        ("White/Grey rows",   f"Matched — high confidence (score ≥ {LOW_CONF_THRESHOLD})"),
        ("Yellow background", f"Matched — low confidence (score < {LOW_CONF_THRESHOLD}) — review recommended"),
        ("Red background",    f"No match found (score < {THRESHOLD}) — left blank"),
    ]
    fills = [None, fill_matched_odd, fill_low_conf, fill_unmatched]
    for r, ((k, v), fill) in enumerate(zip(legend_rows, fills), 1):
        c1 = ws2.cell(row=r, column=1, value=k)
        c2 = ws2.cell(row=r, column=2, value=v)
        for c in (c1, c2):
            c.font = Font(name='Arial', bold=(r == 1), size=10)
        if fill:
            c1.fill = fill

    total     = len(df)
    matched   = (df['2019 Column Name'] != '').sum()
    unmatched = total - matched
    low_conf  = ((df['2019 Column Name'] != '') & (df['Match Score'] < LOW_CONF_THRESHOLD)).sum()

    summary = [
        ("Total 2018 columns:", total),
        ("Matched:",            matched),
        ("Unmatched (blank):",  unmatched),
        ("Low confidence:",     low_conf),
    ]
    ws2.cell(row=6, column=1, value="Summary").font = Font(name='Arial', bold=True, size=10)
    for r, (label, val) in enumerate(summary, 7):
        ws2.cell(row=r, column=1, value=label).font = Font(name='Arial', size=10)
        ws2.cell(row=r, column=2, value=val).font   = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 60

    wb.save(out_path)
    print(f"Saved → {out_path}")
    print(f"  Total: {total}  |  Matched: {matched}  |  Unmatched: {unmatched}  |  Low confidence: {low_conf}")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Map LL87 column names from 2018 to 2019.")
    parser.add_argument('--old', default='LL87_Header_2018.xlsx', help='Path to 2018 header file')
    parser.add_argument('--new', default='LL87_Header_2019.xlsx', help='Path to 2019 header file')
    parser.add_argument('--out', default='LL87_Column_Mapping_2018_to_2019.xlsx', help='Output file path')
    parser.add_argument('--threshold',     type=float, default=THRESHOLD,          help='Min score to count as a match')
    parser.add_argument('--low-conf',      type=float, default=LOW_CONF_THRESHOLD, help='Score below which match is flagged low confidence')
    args = parser.parse_args()

    global THRESHOLD, LOW_CONF_THRESHOLD
    THRESHOLD          = args.threshold
    LOW_CONF_THRESHOLD = args.low_conf

    df = build_mapping(args.old, args.new)
    write_excel(df, args.out)


if __name__ == '__main__':
    main()
